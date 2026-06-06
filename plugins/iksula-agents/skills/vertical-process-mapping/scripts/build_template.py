#!/usr/bin/env python3
"""
Annotated scaffold for building a vertical process-mapping workbook.
Adapt this per vertical: set VERTICAL/ARCHETYPE/OUT, then fill the data lists.
Read references/excel-build-spec.md for the full column/formula/formatting spec.
Helper functions + the styling palette below are the same ones used to build the
gold-standard assets/B2B Distribution.xlsx — reuse them verbatim.

Recommended flow: build in stages (tabs 1-2, then 3-4, then 5-6), each stage
load_workbook + delete-if-exists for idempotency, then run recalc.py and confirm
ZERO formula errors. See assets/B2B Distribution.xlsx for the finished target.

Tab 6 (Iksula Solution Shortlist) carries an Economic Buyer column — derive it with
`from economic_buyer import economic_buyer, BUYER_FILL` (see scripts/economic_buyer.py
for the canonical mapping and the exact 12-column Tab-6 build snippet).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

# ---- EDIT THESE PER VERTICAL ----
VERTICAL = "B2B Distribution"
ARCHETYPE = "Single distributor running its own P&L"
OUT = f"/Users/dj/Iksula FY27/Iksula_Strategy/Services/Vertical - process mapping/{VERTICAL}.xlsx"
# (In the Linux sandbox the same path is /sessions/<id>/mnt/Services/Vertical - process mapping/<VERTICAL>.xlsx)

# ---- PALETTE (keep identical across verticals) ----
FONT="Arial"; NAVY="1F3864"; BLUE="2E5496"; LIGHT="D9E1F2"; LIGHTER="EEF3FB"; GREY="595959"; WHITE="FFFFFF"
GRP_FILL={"Commercial / Revenue":"DDEBF7","Supply Chain / Ops":"E2EFDA","Finance":"FCE4D6","Enabling / Support":"FFF2CC"}
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def hcell(c,t,fill=BLUE,color=WHITE,size=10,align="left"):
    c.value=t; c.font=Font(name=FONT,bold=True,color=color,size=size)
    c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True); c.border=border

def cell(c,t,size=9,align="left",bold=False,color="000000",fill=None,wrap=True):
    c.value=t; c.font=Font(name=FONT,size=size,bold=bold,color=color)
    c.alignment=Alignment(horizontal=align,vertical="top",wrap_text=wrap); c.border=border
    if fill: c.fill=PatternFill("solid",fgColor=fill)

# =====================================================================
# TAB 4 — TASK MATRIX is the heart of the file. Each tuple is ONE L3 task:
# (Group, Function, L1, L2, L3 task, Tools, Pain, Opportunity, Obvious/Non-obvious,
#  Archetype, Value Lever, Driver Metric, Impact, DataReadiness, IntegrationEase,
#  AccuracyTolerance, Reusability, Productizable, Economic Buyer)
# Scores are 1-5 ints. Feasibility/Composite/Tier/QuickWin are FORMULAS (added below).
# Aim for ~100-120 rows across all functions in the taxonomy.
# =====================================================================
TASK_ROWS = [
 ("Finance","Reconciliation & Claims","Scheme Mgmt","Leakage","Detect unclaimed eligible schemes",
  "ERP/Excel","Money left on table","Eligibility vs claimed gap analysis","Non-obvious",
  "NL-Analytics","Leakage Recovery","# schemes x sales", 5,3,3,3, "High","Productizable","CFO"),
 # ... add every L3 task here ...
]

def build_task_matrix(wb):
    for n in ("Task Matrix","Task Matrix1"):
        if n in wb.sheetnames: del wb[n]
    wt=wb.create_sheet("Task Matrix"); wt.sheet_view.showGridLines=False
    TH=["Group","Function","L1 Sub-process","L2 Process","L3 Task","Current Tools","Manual Hotspot / Pain",
        "AI / Automation Opportunity","Obvious / Non-obvious","Solution Archetype","Value Lever",
        "Driver Metric (scales with)","Impact","Data Readiness","Integration Ease","Accuracy Tolerance",
        "Feasibility","Composite Priority","Priority Tier","Quick Win","Reusability","Productizable","Economic Buyer"]
    TW=[16,22,20,20,34,22,34,40,13,20,15,26,8,9,9,10,10,11,9,9,11,13,22]
    for i,(h,w) in enumerate(zip(TH,TW),1):
        hcell(wt.cell(1,i),h,fill=NAVY,align="center"); wt.column_dimensions[get_column_letter(i)].width=w
    wt.freeze_panes="E2"; wt.row_dimensions[1].height=40
    r=2
    for rec in TASK_ROWS:
        (grp,fn,l1,l2,l3,tools,pain,opp,obv,arch,lever,driver,I,D,Ig,A,reuse,prod,buyer)=rec
        f=GRP_FILL.get(grp,LIGHTER)
        cell(wt.cell(r,1),grp,fill=f); cell(wt.cell(r,2),fn,bold=True,fill=f)
        cell(wt.cell(r,3),l1); cell(wt.cell(r,4),l2); cell(wt.cell(r,5),l3,bold=True)
        cell(wt.cell(r,6),tools); cell(wt.cell(r,7),pain); cell(wt.cell(r,8),opp)
        cell(wt.cell(r,9),obv,align="center",color=("C00000" if obv=="Non-obvious" else GREY),bold=(obv=="Non-obvious"))
        cell(wt.cell(r,10),arch,align="center"); cell(wt.cell(r,11),lever,align="center"); cell(wt.cell(r,12),driver)
        for col,val in ((13,I),(14,D),(15,Ig),(16,A)):
            c=wt.cell(r,col); c.value=val; c.font=Font(name=FONT,size=10)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
        wt.cell(r,17).value=f"=ROUND(AVERAGE(N{r}:P{r}),1)"
        wt.cell(r,18).value=f"=ROUND((M{r}+Q{r})/2,1)"
        wt.cell(r,19).value=f'=IF(R{r}>=3.8,"P1",IF(R{r}>=3,"P2","P3"))'
        wt.cell(r,20).value=f'=IF(AND(M{r}>=4,Q{r}>=3.5),"Yes","No")'
        for col in (17,18,19,20):
            c=wt.cell(r,col); c.font=Font(name=FONT,size=10,bold=(col in(18,19)))
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
        cell(wt.cell(r,21),reuse,align="center",bold=(reuse=="High"),color=("006100" if reuse=="High" else "000000"))
        cell(wt.cell(r,22),prod,align="center"); cell(wt.cell(r,23),buyer)
        wt.row_dimensions[r].height=42; r+=1
    last=r-1
    wt.auto_filter.ref=f"A1:W{last}"
    wt.conditional_formatting.add(f"R2:R{last}",ColorScaleRule(start_type='num',start_value=1,start_color='F8696B',
        mid_type='num',mid_value=3,mid_color='FFEB84',end_type='num',end_value=5,end_color='63BE7B'))
    wt.conditional_formatting.add(f"S2:S{last}",CellIsRule(operator='equal',formula=['"P1"'],
        fill=PatternFill('solid',fgColor='C6EFCE'),font=Font(name=FONT,size=10,bold=True,color='006100')))
    wt.conditional_formatting.add(f"T2:T{last}",CellIsRule(operator='equal',formula=['"Yes"'],
        fill=PatternFill('solid',fgColor='FFEB9C'),font=Font(name=FONT,size=10,bold=True)))
    return last

if __name__ == "__main__":
    # Minimal demo build of just the Task Matrix. For a real vertical, also build
    # tabs 1,2,3,5,6 per excel-build-spec.md (mirror the B2B build scripts).
    wb=openpyxl.Workbook()
    n=build_task_matrix(wb)
    wb.save(OUT)
    print(f"Task Matrix built with {n-1} rows -> {OUT}")
    print("Now build the other 5 tabs, then run: python scripts/recalc.py '<file>'")
